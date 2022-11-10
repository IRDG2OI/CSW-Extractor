#####################################
# Transform sheet links to raw urls #
#####################################
import openpyxl

wb = openpyxl.load_workbook('Liens donnees a moissonner G2OI.xlsx')
ws = wb['Geosur']
ws = wb['SEAS-OI']
# ws = wb['Marbec tools COI']
# ws = wb['Ocean Metiss']
# ws = wb['Maurice']
# ws = wb['Comores']

#print(ws.cell(row=1, column=1).value)
# ws.cell(row=2, column=1).hyperlink.target
# ws.cell(row=2, column=1).hyperlink.location

begrow = 1
endrow = ws.max_row

for row in ws.iter_rows(begrow, endrow):
    for cell in row:
        if cell.hyperlink is not None:
            link = cell.hyperlink.target
            location = cell.hyperlink.location
            if location is None:
                location = ''
            else:
                location = '#' + location
            print('"' + cell.value + '"' + ';' 
            + link 
            + location)
        else:
            if cell.value is not None:
                print('"' + cell.value + '"')

 